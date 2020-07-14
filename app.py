import openpyxl as xl
from openpyxl.chart import BarChart,Reference

wb=xl.load_workbook("transactions.xlsx")
sayfa=wb["Sayfa1"]

for satir in range(2,sayfa.max_row+1):
    cell=sayfa.cell(satir, 3)
    hesap= cell.value * 0.9
    hucre_hesap=sayfa.cell(satir,4)
    hucre_hesap.value=hesap

hucre_degerleri=Reference(sayfa,
                          min_row=2,
                          max_row=sayfa.max_row,
                          min_col=4,
                          max_col=4)

chart=BarChart()
chart.add_data(hucre_degerleri)
sayfa.add_chart(chart,"F2")
wb.save("transactions2.xlsx")

